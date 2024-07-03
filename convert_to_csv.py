import argparse

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet


def save_as_csv(excel_file: str, csv_file: str) -> None:
    """
    Saves the Excel file as a CSV file.

    Args:
        excel_file: The path to the Excel file.
        csv_file: The path to save the CSV file.

    Returns:
        None
    """
    wb = load_workbook(excel_file)
    sheet: Worksheet = wb.active
    with open(csv_file, 'w', newline="") as f:
        for row in sheet.iter_rows(values_only=True):
            f.write(";".join([str(cell) if cell is not None else '' for cell in row]) + "\n")


def main(excel_file=None):
    if excel_file is None:
        parser = argparse.ArgumentParser(description='Convert Excel to CSV')
        parser.add_argument('excel_file', help='path to excel file')
        args = parser.parse_args()
        excel_file = args.excel_file

    try:
        save_as_csv(excel_file, excel_file.replace('.xlsx', '.csv'))
    except InvalidFileException as e:
        print(e)


if __name__ == '__main__':
    main()
